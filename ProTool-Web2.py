import os
import json
import base64
import io
import re
import streamlit as st
import tempfile
from openai import OpenAI
import fitz  # PyMuPDF
import openpyxl
from datetime import datetime
import requests
from pathlib import Path

# Tratamento seguro caso o arquivo busca_cad.py não esteja no mesmo diretório durante testes
try:
    from busca_cad import BuscaTripla
except ImportError:
    st.warning("⚠️ Módulo 'busca_cad' não encontrado no diretório. Usando classe simulada para contingência.")
    class BuscaTripla:
        def iniciar(self, path): pass
        def busca(self, rua, lote, quadra): return {'sucesso': False, 'zona_final': 'ZH2'}

# --- CONFIGURAÇÃO INICIAL E MODELO ---
MODEL = "gpt-5.4-mini" # Recomendo manter um modelo válido atualizado
MODELO_ARQUIVO = "MODELO_Memorial_Planilha_Declaração_R00.xlsx"
CAMINHO_MAPA_DXF = "mapa_zoneamento4.dxf"

@st.cache_resource(show_spinner="Carregando mapa cadastral na memória do servidor...")
def inicializar_motor_busca():
    """
    Carrega o arquivo DXF e inicializa o motor geométrico apenas uma vez.
    O resultado fica salvo na memória RAM (Cache) do Streamlit Cloud.
    """
    buscador = BuscaTripla()
    
    # Verifica se o arquivo existe no repositório
    if os.path.exists(CAMINHO_MAPA_DXF):
        sucesso = buscador.iniciar(CAMINHO_MAPA_DXF)
        return buscador, True
    else:
        return buscador, False

# Executa a função (ela só roda de fato na primeira vez que o app liga)
motor_busca, mapa_disponivel = inicializar_motor_busca()

# PASSO 2: Integração e cruzamento geoespacial via mapa DXF em Memória
if mapa_disponivel:
                st.info("📍 Cruzando coordenadas no mapa cadastral em memória...")
                
                rua_alvo = res_mat.get('confrontacao_frente')
                lote_alvo = res_mat.get('lote')
                quadra_alva = res_mat.get('quadra')
                
                # O motor_busca já está pronto e carregado do cache!
                res_zona = motor_busca.busca(rua_alvo, lote_alvo, quadra_alva)
                
                if res_zona.get('sucesso'):
                    st.session_state['zona_final'] = res_zona['zona_final']
                    st.success(f"✅ Lote localizado! Zona encontrada: {res_zona['zona_final']}")
                else:
                    st.warning("⚠️ Lote não localizado geometricamente no DXF. Definida zona padrão: ZH2.")
                    st.session_state['zona_final'] = "ZH2"
else:
                st.warning("⚠️ Arquivo DXF interno não encontrado no servidor. Definida zona padrão: ZH2.")
                st.session_state['zona_final'] = "ZH2"


st.set_page_config(page_title="Análise Documental - Prefeituras", layout="wide")

# Inicialização do Session State para persistência
if 'dados_extraidos' not in st.session_state:
    st.session_state['dados_extraidos'] = None
if 'arquivo_gerado' not in st.session_state:
    st.session_state['arquivo_gerado'] = None
if 'nome_proprietario' not in st.session_state:
    st.session_state['nome_proprietario'] = "DOC"
if 'pasta_origem' not in st.session_state:
    st.session_state['pasta_origem'] = None
if 'caminho_arquivo_salvo' not in st.session_state:
    st.session_state['caminho_arquivo_salvo'] = None
if 'cache_ceps' not in st.session_state:
    st.session_state['cache_ceps'] = {}
if 'tags_debug' not in st.session_state:
    st.session_state['tags_debug'] = {}
if 'zona_final' not in st.session_state:
    st.session_state['zona_final'] = None

# --- FUNÇÕES AUXILIARES DE LIMPEZA E FORMATAÇÃO ---

def formatar_numero_br(valor):
    """
    Remove letras/unidades e garante que o separador decimal seja estritamente a VÍRGULA.
    Ex: "1.234.567,89 m²" -> "1234567,89"
    Ex: "1234.56" -> "1234,56"
    """
    if valor is None:
        return ""
    
    valor_str = str(valor).strip()
    
    # Remove tudo que não for dígito, ponto ou vírgula (ex: m, m², %, letras)
    apenas_numeros = re.sub(r'[^\d.,]', '', valor_str)
    
    if not apenas_numeros:
        return ""

    qtd_pontos = apenas_numeros.count('.')
    qtd_virgulas = apenas_numeros.count(',')

    # Cenário 1: Tem ponto e vírgula (ex: 1.234,56 ou 1,234.56)
    if qtd_pontos > 0 and qtd_virgulas > 0:
        pos_ponto = apenas_numeros.rfind('.')
        pos_virgula = apenas_numeros.rfind(',')
        if pos_ponto > pos_virgula:
            # Formato US (1,234.56) -> Remove vírgula, troca ponto por vírgula
            apenas_numeros = apenas_numeros.replace(',', '').replace('.', ',')
        else:
            # Formato BR (1.234,56) -> Remove o ponto de milhar
            apenas_numeros = apenas_numeros.replace('.', '')
            
    # Cenário 2: Só tem ponto (ex: 1234.56 ou 1.234.567)
    elif qtd_pontos > 0:
        # Se tem mais de um ponto ou o ponto não isola os 2/3 últimos dígitos, é separador de milhar
        if qtd_pontos > 1 or (len(apenas_numeros) - apenas_numeros.rfind('.') > 3 and len(apenas_numeros) > 4):
            apenas_numeros = apenas_numeros.replace('.', '')
        else:
            # Se for casa decimal, converte o ponto para vírgula
            apenas_numeros = apenas_numeros.replace('.', ',')
            
    # Cenário 3: Só tem vírgula, mantemos a última se houver mais de uma
    elif qtd_virgulas > 1:
        partes = apenas_numeros.rsplit(',', 1)
        apenas_numeros = partes[0].replace(',', '') + ',' + partes[1]

    return apenas_numeros

def formatar_documento_detalhado(valor):
    if valor is None or str(valor).strip() == "":
        return "CPF/CNPJ:", ""
    numeros = re.sub(r'\D', '', str(valor))
    if not numeros:
        return "CPF/CNPJ:", str(valor)
    if len(numeros) <= 11:
        numeros = numeros.zfill(11)
        return "CPF:", f"{numeros[:3]}.{numeros[3:6]}.{numeros[6:9]}-{numeros[9:]}"
    else:
        numeros = numeros.zfill(14)
        return "CNPJ:", f"{numeros[:2]}.{numeros[2:5]}.{numeros[5:8]}/{numeros[8:12]}-{numeros[12:]}"

def formatar_cep(cep_bruto):
    if not cep_bruto:
        return ""
    cep_limpo = re.sub(r'\D', '', str(cep_bruto))
    if len(cep_limpo) != 8:
        return cep_bruto
    return f"{cep_limpo[:5]}-{cep_limpo[5:]}"

def extrair_termos_significativos(texto):
    termos_ignorar = [
        'bairro', 'loteamento', 'jardim', 'jd', 'residencial', 'comercial',
        'industrial', 'parque', 'condomínio', 'cond', 'avenida', 'av', 'rua',
        'r', 'alameda', 'travessa', 'pça', 'praça', 'rodovia', 'estrada',
        'via', 'trav', 'estr', 'rod', 'de', 'do', 'da', 'dos', 'das',
        'e', 'ou', 'norte', 'sul', 'leste', 'oeste', 'n', 's', 'l', 'o'
    ]
    if not texto:
        return []
    texto_limpo = re.sub(r'\d+', '', str(texto)).strip()
    texto_limpo = re.sub(r'[^\w\s]', ' ', texto_limpo)
    texto_limpo = re.sub(r'\s+', ' ', texto_limpo).strip()
    palavras = texto_limpo.upper().split()
    return [p for p in palavras if p.lower() not in termos_ignorar]

def selecionar_cep_por_loteamento(resultados_cep, loteamento):
    if not resultados_cep:
        return None
    if len(resultados_cep) == 1:
        return resultados_cep[0]
    
    termos_lote = set(extrair_termos_significativos(loteamento))
    st.info(f"🔍 Comparando {len(resultados_cep)} resultados de CEP com loteamento: **{loteamento}**")
    
    with st.expander("📊 Detalhes da Comparação de Loteamentos", expanded=False):
        st.write(f"**Termos significativos do loteamento:** {termos_lote if termos_lote else 'Nenhum'}")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.write("**Resultado**")
            for idx in range(len(resultados_cep)):
                st.write(f"`Opção {idx + 1}`")
        with col2:
            st.write("**Bairro**")
            for resultado in resultados_cep:
                st.write(f"{resultado.get('bairro', 'N/A')}")
        with col3:
            st.write("**Coincidências**")
            for resultado in resultados_cep:
                termos_bairro = set(extrair_termos_significativos(resultado.get('bairro', '')))
                coincidencias = termos_lote.intersection(termos_bairro)
                st.write(f"{len(coincidencias)} termos")
    
    pontuacoes = []
    for idx, resultado in enumerate(resultados_cep):
        termos_bairro = set(extrair_termos_significativos(resultado.get('bairro', '')))
        coincidencias = termos_lote.intersection(termos_bairro)
        pontuacao = len(coincidencias)
        pontuacoes.append({
            'indice': idx,
            'resultado': resultado,
            'pontuacao': pontuacao,
            'coincidencias': coincidencias
        })
    
    pontuacoes_ordenadas = sorted(pontuacoes, key=lambda x: x['pontuacao'], reverse=True)
    
    if pontuacoes_ordenadas[0]['pontuacao'] > 0:
        resultado_selecionado = pontuacoes_ordenadas[0]
        cep_formatado = formatar_cep(resultado_selecionado['resultado'].get('cep'))
        st.success(f"✅ CEP selecionado (Opção {resultado_selecionado['indice'] + 1}): `{cep_formatado}` - **{resultado_selecionado['resultado'].get('bairro')}** ({resultado_selecionado['pontuacao']} coincidências)")
        return resultado_selecionado['resultado']
    else:
        cep_formatado = formatar_cep(resultados_cep[0].get('cep'))
        st.warning(f"⚠️ Nenhuma correspondência encontrada. Usando primeiro resultado: `{cep_formatado}` - **{resultados_cep[0].get('bairro')}**")
        return resultados_cep[0]

def limpar_rua_para_cep(rua):
    termos_ignorar = [
        'norte', 'sul', 'leste', 'oeste', 'n', 's', 'l', 'o',
        'av', 'avenida', 'rua', 'r', 'alameda', 'travessa', 'praca', 'praça',
        'rod', 'rodovia', 'estr', 'estrada', 'via', 'pça', 'trav'
    ]
    if not rua:
        return ""
    rua_limpa = re.sub(r'\d+', '', str(rua)).strip()
    rua_limpa = re.sub(r'[^\w\s]', ' ', rua_limpa)
    rua_limpa = re.sub(r'\s+', ' ', rua_limpa).strip()
    palavras = rua_limpa.upper().split()
    if not palavras:
        return ""
    
    palavras_significativas = []
    for palavra in reversed(palavras):
        if palavra.lower() not in termos_ignorar:
            palavras_significativas.append(palavra)
        elif palavras_significativas:
            break
            
    palavras_significativas.reverse()
    resultado = ' '.join(palavras_significativas)
    return resultado if resultado else rua_limpa

def buscar_cep_viacep(rua_limpa, cidade_limpa, estado_limpo):
    try:
        url_viacep = f"https://viacep.com.br/ws/{estado_limpo}/{cidade_limpa}/{rua_limpa}/json/"
        proxies = [
            ("allorigins", "https://api.allorigins.win/raw"),
            ("corsfix", "https://cors-anywhere.herokuapp.com/"),
        ]
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
        
        for proxy_nome, proxy_url in proxies:
            try:
                if "allorigins" in proxy_url:
                    response = requests.get(proxy_url, params={"url": url_viacep}, headers=headers, timeout=10)
                else:
                    response = requests.get(proxy_url + url_viacep, headers=headers, timeout=10)
                
                if response.status_code == 200:
                    dados = response.json()
                    if isinstance(dados, (list, dict)) and not (isinstance(dados, dict) and 'erro' in dados):
                        st.info(f"✅ CEP encontrado via ViaCEP ({proxy_nome})")
                        return dados
            except:
                continue
        
        try:
            response = requests.get(url_viacep, headers=headers, timeout=10)
            if response.status_code == 200:
                dados = response.json()
                if isinstance(dados, (list, dict)) and not (isinstance(dados, dict) and 'erro' in dados):
                    st.info("✅ CEP encontrado via ViaCEP (conexão direta)")
                    return dados
        except:
            pass
        return None
    except Exception as e:
        st.warning(f"⚠️ Erro ao buscar via ViaCEP: {str(e)}")
        return None

def buscar_cep_brasilapi(rua_limpa, cidade_limpa, estado_limpo):
    try:
        url_brasilapi = f"https://brasilapi.com.br/api/address/v2/{estado_limpo}/{cidade_limpa}/{rua_limpa}"
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
        response = requests.get(url_brasilapi, headers=headers, timeout=10)
        
        if response.status_code == 200:
            dados = response.json()
            if isinstance(dados, list) and len(dados) > 0:
                st.info("✅ CEP encontrado via BrasilAPI")
                return dados
            elif isinstance(dados, dict) and 'bairro' in dados:
                st.info("✅ CEP encontrado via BrasilAPI")
                return [dados]
        return None
    except Exception as e:
        st.warning(f"⚠️ Erro ao buscar via BrasilAPI: {str(e)}")
        return None

def buscar_cep_por_endereco(rua, cidade, estado, loteamento=""):
    try:
        if not rua or not cidade or not estado:
            st.warning("⚠️ Dados insuficientes para buscar CEP (rua, cidade ou estado vazio)")
            return ""
        
        chave_cache = f"{rua}|{cidade}|{estado}".upper()
        if chave_cache in st.session_state['cache_ceps']:
            st.info(f"📦 CEP obtido do cache")
            return st.session_state['cache_ceps'][chave_cache]
        
        rua_limpa = limpar_rua_para_cep(rua)
        cidade_limpa = str(cidade).strip().upper()
        estado_limpo = str(estado).strip().upper()[:2]
        
        if not rua_limpa or len(rua_limpa) < 3:
            st.warning(f"⚠️ Nome da rua muito curto após limpeza: '{rua_limpa}'")
            return ""
        
        with st.expander("🔍 Detalhes da Busca de CEP", expanded=False):
            col1, col2 = st.columns(2)
            with col1:
                st.write("**Dados Originais:**")
                st.write(f"- Rua: `{rua}` | Cidade: `{cidade}` | Estado: `{estado}` | Loteamento: `{loteamento}`")
            with col2:
                st.write("**Dados Processados para API:**")
                st.write(f"- Rua (limpa): `{rua_limpa}` | Cidade: `{cidade_limpa}` | Estado: `{estado_limpo}`")
        
        st.info("🌐 Buscando CEP via ViaCEP...")
        dados = buscar_cep_viacep(rua_limpa, cidade_limpa, estado_limpo)
        
        if dados is None:
            st.warning("⚠️ ViaCEP indisponível, tentando BrasilAPI...")
            dados = buscar_cep_brasilapi(rua_limpa, cidade_limpa, estado_limpo)
        
        if dados is None:
            st.error(f"❌ CEP não encontrado para: {rua_limpa}, {cidade_limpa} - {estado_limpo}")
            return ""
        
        if isinstance(dados, list) and len(dados) > 0:
            st.success(f"✅ {len(dados)} resultado(s) encontrado(s)")
            resultado_selecionado = selecionar_cep_por_loteamento(dados, loteamento) if len(dados) > 1 and loteamento else dados[0]
            cep = resultado_selecionado.get('cep', '')
            if cep:
                cep_formatado = formatar_cep(cep)
                st.success(f"✅ CEP selecionado: {cep_formatado}")
                st.session_state['cache_ceps'][chave_cache] = cep_formatado
                return cep_formatado
        return ""
    except Exception as e:
        st.error(f"⚠️ Erro inesperado ao buscar CEP: {str(e)}")
        return ""

def obter_pasta_origem(uploaded_file):
    try:
        home = str(Path.home())
        desktop = Path(home) / "Desktop"
        if desktop.exists(): 
            return str(desktop)
        return home
    except:
        return None

def salvar_arquivo_xlsx(arquivo_bytes, nome_arquivo, pasta_destino):
    try:
        pasta_path = Path(pasta_destino)
        pasta_path.mkdir(parents=True, exist_ok=True)
        caminho_completo = pasta_path / nome_arquivo
        with open(caminho_completo, 'wb') as f:
            f.write(arquivo_bytes.getvalue())
        return str(caminho_completo)
    except Exception as e:
        st.error(f"Erro ao salvar arquivo localmente: {e}")
        return None

def extract_data_from_document(prompt_text, uploaded_file, api_key):
    client = OpenAI(api_key=api_key)
    user_content = [{"type": "text", "text": prompt_text}]
    file_bytes = uploaded_file.read()
    uploaded_file.seek(0)
    file_ext = uploaded_file.name.split('.')[-1].lower()

    if file_ext in ['png', 'jpg', 'jpeg']:
        b64 = base64.b64encode(file_bytes).decode('utf-8')
        user_content.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}})
    elif file_ext == 'pdf':
        try:
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            for i, page in enumerate(doc):
                if i >= 5: break
                pix = page.get_pixmap(dpi=150)
                b64 = base64.b64encode(pix.tobytes("jpeg")).decode('utf-8')
                user_content.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}})
            doc.close()
        except Exception as e:
            raise ValueError(f"Erro ao processar o PDF: {e}")

    response = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": "Retorne estritamente JSON válido."},
            {"role": "user", "content": user_content}
        ],
        response_format={"type": "json_object"},
        temperature=0.1
    )
    return json.loads(response.choices[0].message.content)

def extract_data_from_multiple_files(prompt_text, uploaded_files, api_key):
    client = OpenAI(api_key=api_key)
    user_content = [{"type": "text", "text": prompt_text}]
    
    for f in uploaded_files:
        f_bytes = f.read()
        f.seek(0)
        ext = f.name.split('.')[-1].lower()
        if ext in ['png', 'jpg', 'jpeg']:
            b64 = base64.b64encode(f_bytes).decode('utf-8')
            user_content.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}})
        elif ext == 'pdf':
            try:
                doc = fitz.open(stream=f_bytes, filetype="pdf")
                for i, page in enumerate(doc):
                    if i >= 5: break
                    pix = page.get_pixmap(dpi=150)
                    b64 = base64.b64encode(pix.tobytes("jpeg")).decode('utf-8')
                    user_content.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}})
                doc.close()
            except Exception as e:
                raise ValueError(f"Erro ao processar o PDF do Projeto: {e}")
    
    response = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": "Retorne estritamente JSON válido."},
            {"role": "user", "content": user_content}
        ],
        response_format={"type": "json_object"},
        temperature=0.1
    )
    return json.loads(response.choices[0].message.content)

def preencher_aba_com_tags(ws, contexto):
    tags_encontradas = {}
    tags_substituidas = {}
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for k, v in contexto.items():
                    tag = "{{" + k + "}}"
                    if tag in cell.value:
                        texto_substituto = str(v) if v is not None else ""
                        
                        if tag not in tags_encontradas:
                            tags_encontradas[tag] = []
                        tags_encontradas[tag].append({
                            'celula': cell.coordinate,
                            'conteudo_original': cell.value,
                            'valor': texto_substituto
                        })
                        
                        cell.value = cell.value.replace(tag, texto_substituto)
                        
                        if tag not in tags_substituidas:
                            tags_substituidas[tag] = 0
                        tags_substituidas[tag] += 1
    return tags_encontradas, tags_substituidas

def criar_contexto_dados(dados, zona_dxf=None):
    mat = dados.get('matricula', {})
    idf1 = dados.get('identificacao_1', {})
    idf2 = dados.get('identificacao_2', {})
    prj = dados.get('projeto', {})
    
    tipo_doc_1, num_doc_1 = formatar_documento_detalhado(idf1.get('cnpj_cpf'))
    tipo_doc_2, num_doc_2 = formatar_documento_detalhado(idf2.get('cnpj_cpf')) if idf2.get('cnpj_cpf') else ("", "")
    
    agora = datetime.now()
    meses = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
    
    rua = str(mat.get('confrontacao_frente', '')).strip()
    cidade = str(mat.get('cidade', '')).strip().upper()
    estado = str(mat.get('estado', '')).strip().upper()
    loteamento = str(mat.get('loteamento', '')).strip()
    
    cep = buscar_cep_por_endereco(rua, cidade, estado, loteamento) if rua and cidade and estado else ""
    
    contexto = {
        'cnm': mat.get('cnm', ''),
        'matricula': mat.get('matricula', ''),
        'folha': mat.get('folha', ''),
        'cartorio': mat.get('cartorio', ''),
        'livro': mat.get('livro', ''),
        'data_registro': mat.get('data_registro', ''),
        'data_documento': mat.get('data_documento', ''),
        'loteamento': loteamento.upper(),
        'lote': str(mat.get('lote', '')).upper(),
        'quadra': str(mat.get('quadra', '')).upper(),
        'cidade': cidade,
        'estado': estado,
        
        # --- DADOS NUMÉRICOS EXTREMAMENTE FORMATADOS PARA EXCEL ---
        'area': formatar_numero_br(mat.get('area', '')),
        'area_construida_total': formatar_numero_br(prj.get('area_construida_total', '')),
        'numero_pavimentos': formatar_numero_br(prj.get('numero_pavimentos', '')),
        'numero_vagas': formatar_numero_br(prj.get('numero_vagas', '')),
        'inclinacao_telhado': formatar_numero_br(prj.get('inclinacao_telhado', '')),
        'altura_maxima': formatar_numero_br(prj.get('altura_maxima', '')),
        
        'confrontacao_frente': rua.upper(),
        'confrontacao_fundos': str(mat.get('confrontacao_fundos', '')).upper(),
        'confrontacao_lado_direito': str(mat.get('confrontacao_lado_direito', '')).upper(),
        'confrontacao_lado_esquerdo': str(mat.get('confrontacao_lado_esquerdo', '')).upper(),

        # Proprietário 1
        'proprietario_1': str(idf1.get('proprietario', '')).upper(),
        'tipo_doc_1': tipo_doc_1.replace(':', ''),
        'num_doc_1': num_doc_1,
        'doc_completo_1': f"{tipo_doc_1} {num_doc_1}" if num_doc_1 else "",

        # Proprietário 2
        'proprietario_2': str(idf2.get('proprietario', '')).upper() if idf2.get('proprietario') else "",
        'tipo_doc_2': tipo_doc_2.replace(':', ''),
        'num_doc_2': num_doc_2,
        'doc_completo_2': f"{tipo_doc_2} {num_doc_2}" if num_doc_2 else "",

        'finalidade_obra': str(prj.get('finalidade_obra', '')).upper(),
        'desenhista': str(prj.get('desenhista', '')).upper(),
        'tipo_telhado': str(prj.get('tipo_telhado', '')).upper(),
        'tipo_forro': str(prj.get('tipo_forro', '')).upper(),
        'endereco_obra': str(prj.get('endereco_obra', '')).upper(),
        'cep': cep,
        'zona_zoneamento': zona_dxf if zona_dxf else "ZH2",
        
        'data_atual': agora.strftime("%d/%m/%Y"),
        'data_extensa': f"Sorriso - MT, {agora.day:02d} de {meses[agora.month]} de {agora.year}"
    }
    return contexto

def gerar_arquivo_final(dados, zona_dxf=None):
    contexto = criar_contexto_dados(dados, zona_dxf)
    proprietario_1 = contexto['proprietario_1']
    
    try:
        with st.expander("🔍 DEBUG - Contexto Carregado", expanded=False):
            st.json({
                'proprietario_1': contexto.get('proprietario_1'),
                'cep': contexto.get('cep'),
                'zona_zoneamento': contexto.get('zona_zoneamento'),
                'loteamento': contexto.get('loteamento'),
                'area_formatada': contexto.get('area'),
                'altura_formatada': contexto.get('altura_maxima')
            })
        
        wb_completo = openpyxl.load_workbook(MODELO_ARQUIVO)
        todas_as_tags = {}
        
        for nome_aba in wb_completo.sheetnames:
            ws = wb_completo[nome_aba]
            tags_encontradas, tags_substituidas = preencher_aba_com_tags(ws, contexto)
            if tags_encontradas or tags_substituidas:
                todas_as_tags[nome_aba] = {
                    'encontradas': tags_encontradas,
                    'substituidas': tags_substituidas
                }
        
        with st.expander("📋 DEBUG - Tags Processadas por Aba", expanded=False):
            if todas_as_tags:
                for aba, dados_aba in todas_as_tags.items():
                    st.write(f"### Aba: **{aba}**")
                    for tag, info in dados_aba['encontradas'].items():
                        st.write(f"#### {tag} - {len(info)} ocorrência(s)")
                        for item in info:
                            st.write(f"- Célula: `{item['celula']}` | Valor Injetado: `{item['valor']}`")
            else:
                st.warning("⚠️ Nenhuma tag detectada ou preenchida. Verifique os delimitadores `{{ }}` no arquivo Excel.")
        
        xlsx_preenchido_io = io.BytesIO()
        wb_completo.save(xlsx_preenchido_io)
        xlsx_preenchido_io.seek(0)
        wb_completo.close()
        
        st.session_state['tags_debug'] = todas_as_tags
        
        return {
            'arquivo_xlsx': xlsx_preenchido_io,
            'nome_proprietario': proprietario_1,
            'tags_processadas': todas_as_tags
        }
    except FileNotFoundError:
        raise FileNotFoundError(f"Arquivo '{MODELO_ARQUIVO}' não encontrado.")
    except Exception as e:
        raise Exception(f"Erro ao gerar arquivo final: {str(e)}")

# --- INTERFACE STREAMLIT ---
st.title("🏛️ Automação Documental - Prefeituras")
st.markdown("*Selecione ou arraste os arquivos para extrair as informações e preencher os templates.*")
st.success("✅ Arquivo será gerado em XLSX e salvo automaticamente na pasta de origem")

api_key = st.text_input("OpenAI API Key (Deixe em branco se configurada no sistema):", type="password")
if not api_key: 
    api_key = os.environ.get("OPENAI_API_KEY")

col1, col2, col3 = st.columns(3)
with col1:
    f_mat = st.file_uploader("📋 Matrícula do Imóvel", type=["pdf", "png", "jpg", "jpeg"], key="matricula")
with col2:
    f_idf_1 = st.file_uploader("🪪 Documento de Identificação - Proprietário 1", type=["pdf", "png", "jpg", "jpeg"], key="idf_1")
with col3:
    f_idf_2 = st.file_uploader("🪪 Documento de Identificação - Proprietário 2 (Opcional)", type=["pdf", "png", "jpg", "jpeg"], key="idf_2")

col_prj, col_dxf = st.columns(2)
with col_prj:
    f_prj = st.file_uploader("📐 Pranchas Arquitetônicas", type=["pdf", "png", "jpg", "jpeg"], accept_multiple_files=True, key="projeto")

if st.button("⚙️ Processar e Gerar Documentos", type="primary", use_container_width=True):
    if not api_key:
        st.error("❌ Por favor, insira a chave da OpenAI para continuar.")
        st.stop()
    
    if not f_mat or not f_idf_1 or not f_prj:
        st.error("❌ Por favor, preencha os campos obrigatórios: Matrícula, Documento Proprietário 1 e Pranchas Arquitetônicas.")
        st.stop()
    
    pasta_origem = obter_pasta_origem(f_mat)
    st.session_state['pasta_origem'] = pasta_origem
        
    prompts = {
        "matricula": (
            "Analise a matrícula do imóvel (terreno) e extraia os seguintes dados em JSON com exatamente estas chaves: "
            "cnm, matricula, folha, cartorio, livro, data_registro, data_documento, "
            "loteamento, lote, quadra, cidade, estado, area (extraia APENAS o valor numérico, utilizando VÍRGULA como separador decimal, sem qualquer unidade de medida), "
            "confrontacao_frente (descrição do limite frontal, ex: 'Av. dos Imigrantes Sul'), "
            "confrontacao_fundos (descrição do limite dos fundos, ex: 'Lote 06 e 07'), "
            "confrontacao_lado_direito (descrição do limite direito, ex: 'Lotes 19 e 20'), "
            "confrontacao_lado_esquerdo (descrição do limite esquerdo, ex: 'Lotes 21 e 22')."
        ),
        "identificacao": (
            "Analise o documento de identificação do proprietário e extraia os seguintes dados em JSON com exatamente estas chaves: "
            "proprietario (nome completo), cnpj_cpf."
        ),
        "projeto": (
            "Analise as pranchas arquitetônicas e extraia os seguintes dados em JSON com exatamente estas chaves: "
            "area_construida_total (APENAS o valor numérico com vírgula como decimal, sem unidades), numero_pavimentos (apenas número inteiro), numero_vagas (apenas número inteiro), "
            "finalidade_obra, desenhista, tipo_telhado, inclinacao_telhado (APENAS número, sem o símbolo %), tipo_forro, "
            "altura_maxima (utilize como referência os desenhos de 'Corte' nas pranchas, extraia APENAS o número utilizando vírgula como decimal), endereco_obra."
        ),
    }

    with st.spinner("⏳ Analisando documentos da prefeitura e preenchendo arquivo... Isso pode levar alguns segundos."):
        try:
            # PASSO 1: Extrai dados dos documentos estruturados por IA
            st.info("📄 Extraindo dados da Matrícula...")
            res_mat = extract_data_from_document(prompts["matricula"], f_mat, api_key)
            
            st.info("🪪 Extraindo dados do Proprietário 1...")
            res_idf_1 = extract_data_from_document(prompts["identificacao"], f_idf_1, api_key)
            
            res_idf_2 = extract_data_from_document(prompts["identificacao"], f_idf_2, api_key) if f_idf_2 else {}
            
            st.info("📐 Extraindo dados das Pranchas Arquitetônicas...")
            res_prj = extract_data_from_multiple_files(prompts["projeto"], f_prj, api_key)
            
            st.session_state['dados_extraidos'] = {
                "matricula": res_mat, 
                "identificacao_1": res_idf_1, 
                "identificacao_2": res_idf_2,
                "projeto": res_prj
            }
            
            # PASSO 2: Integração e cruzamento geoespacial via mapa DXF
            if f_dxf:
                st.info("📍 Localizando lote no mapa cadastral DXF...")
                with tempfile.NamedTemporaryFile(delete=False, suffix='.dxf') as tmp:
                    tmp.write(f_dxf.getvalue())
                    tmp_path = tmp.name
                
                buscador = BuscaTripla()
                buscador.iniciar(tmp_path)
                
                rua_alvo = res_mat.get('confrontacao_frente')
                lote_alvo = res_mat.get('lote')
                quadra_alva = res_mat.get('quadra')
                
                res_zona = buscador.busca(rua_alvo, lote_alvo, quadra_alva)
                os.remove(tmp_path)
                
                if res_zona.get('sucesso'):
                    st.session_state['zona_final'] = res_zona['zona_final']
                    st.success(f"✅ Lote localizado! Zona encontrada: {res_zona['zona_final']}")
                else:
                    st.warning("⚠️ Lote não localizado geometricamente no DXF. Definida zona padrão: ZH2.")
                    st.session_state['zona_final'] = "ZH2"
            else:
                st.session_state['zona_final'] = "ZH2"
            
            # PASSO 3: Geração do arquivo processando as tags e a zona obtida
            st.info("📊 Preenchendo template e gerando matriz XLSX...")
            resultado = gerar_arquivo_final(st.session_state['dados_extraidos'], st.session_state['zona_final'])
            
            st.session_state['arquivo_gerado'] = resultado['arquivo_xlsx']
            st.session_state['nome_proprietario'] = resultado['nome_proprietario'] if resultado['nome_proprietario'] else "DOC"
            
            # PASSO 4: Salvamento Automatizado na pasta alvo
            if pasta_origem:
                nome_arquivo = f"{st.session_state['nome_proprietario']}_Memorial_Planilha_Declaração_R00.xlsx"
                caminho_salvo = salvar_arquivo_xlsx(st.session_state['arquivo_gerado'], nome_arquivo, pasta_origem)
                
                if caminho_salvo:
                    st.session_state['caminho_arquivo_salvo'] = caminho_salvo
                    st.success(f"✅ Arquivo salvo automaticamente em:\n`{caminho_salvo}`")
                else:
                    st.warning("⚠️ Não foi possível salvar o arquivo automaticamente. Faça o download manual abaixo.")
            else:
                st.warning("⚠️ Não foi possível mapear a pasta de origem do arquivo de entrada. Faça o download manual.")
            
            st.success("✅ Análise documental finalizada com sucesso!")
            
        except Exception as e:
            st.error(f"❌ Erro crítico no pipeline: {e}")
            import traceback
            st.error(traceback.format_exc())

# --- EXIBIÇÃO E BOTÕES DE DOWNLOAD ---
if st.session_state['dados_extraidos']:
    st.divider()
    st.subheader("📄 Documento Pronto para Download")
    
    nome_proprietario = st.session_state['nome_proprietario']
    
    if st.session_state['arquivo_gerado']:
        st.download_button(
            label="📥 Baixar XLSX Preenchido",
            data=st.session_state['arquivo_gerado'],
            file_name=f"{nome_proprietario}_Memorial_Planilha_Declaração_R00.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="xlsx_download"
        )
    else:
        st.error("❌ Erro ao expor o arquivo binário gerado.")

    st.divider()
    with st.expander("📋 Ver Dados Brutos Extraídos (JSON)", expanded=False):
        st.json(st.session_state['dados_extraidos'])